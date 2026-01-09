import os
import re
from datetime import datetime, time
from typing import Optional, Dict, Tuple

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak


# =========================================================
# Helpers
# =========================================================
def safe_date(x):
    dt = pd.to_datetime(x, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


def clean_name(nama: str) -> str:
    if nama is None:
        return ""
    nama = str(nama).upper().strip()
    nama = re.sub(r"[^\w\s]", "", nama)
    nama = re.sub(r"\s+", " ", nama)
    return nama


def safe_folder_name(text: str) -> str:
    text = str(text).strip().replace(" ", "_")
    keep = [c for c in text if c.isalnum() or c in ("_", "-", ".")]
    return "".join(keep)[:80]


def parse_time_safe(val) -> Optional[time]:
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, time):
        return val
    dt = pd.to_datetime(val, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.time()


def minutes_diff(t1: Optional[time], t2: Optional[time]) -> Optional[int]:
    if t1 is None or t2 is None:
        return None
    dt1 = datetime.combine(datetime.today(), t1)
    dt2 = datetime.combine(datetime.today(), t2)
    return int((dt2 - dt1).total_seconds() // 60)


# =========================================================
# Manual Override
# =========================================================
MANUAL_ALLOWED = {"S", "I", "C", "DL"}


def apply_manual_override(df: pd.DataFrame, df_manual: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Enterprise Manual Override

    Matching:
      - If override has NIK -> match by (NIK + tanggal)
      - Else -> match by (Nama_clean + tanggal) even if base has NIK
    """
    out = df.copy()

    if "Tanggal" not in out.columns:
        return out
    if "Nama" not in out.columns:
        return out
    if "NIK" not in out.columns:
        out["NIK"] = ""

    out["Tanggal_only"] = out["Tanggal"].apply(safe_date)
    out["Nama_clean"] = out["Nama"].apply(clean_name)
    out["NIK"] = out["NIK"].astype(str).fillna("").str.strip()

    out["Manual Status"] = ""

    if df_manual is None or df_manual.empty:
        return out

    m = df_manual.copy()

    if "Tanggal" not in m.columns or "Nama" not in m.columns or "Status Manual" not in m.columns:
        return out

    if "NIK" not in m.columns:
        m["NIK"] = ""

    m["Tanggal_only"] = m["Tanggal"].apply(safe_date)
    m["Nama_clean"] = m["Nama"].apply(clean_name)
    m["NIK"] = m["NIK"].astype(str).fillna("").str.strip()
    m["Status Manual"] = m["Status Manual"].astype(str).str.upper().str.strip()

    m = m[m["Status Manual"].isin(MANUAL_ALLOWED)]
    m = m.dropna(subset=["Tanggal_only"])
    if m.empty:
        return out

    nik_map = {}
    m_nik = m[m["NIK"] != ""]
    for _, r in m_nik.iterrows():
        key = f"{r['NIK']}|{r['Tanggal_only']}"
        nik_map[key] = r["Status Manual"]

    nama_map = {}
    m_nama = m[m["NIK"] == ""]
    for _, r in m_nama.iterrows():
        key = f"{r['Nama_clean']}|{r['Tanggal_only']}"
        nama_map[key] = r["Status Manual"]

    def resolve(row):
        if row["Tanggal_only"] is None:
            return ""
        key_nik = f"{row['NIK']}|{row['Tanggal_only']}"
        key_nama = f"{row['Nama_clean']}|{row['Tanggal_only']}"
        if key_nik in nik_map:
            return nik_map[key_nik]
        if key_nama in nama_map:
            return nama_map[key_nama]
        return ""

    out["Manual Status"] = out.apply(resolve, axis=1)
    return out


# =========================================================
# Status Logic (Enterprise Policy)
# =========================================================
def compute_daily_status(row, batas_masuk="07:30:00", batas_pulang="16:00:00") -> Tuple[str, int, int]:
    """
    FINAL STATUS POLICY:
    1) Manual Status -> S/I/C/DL
    2) If scan masuk OR scan pulang missing -> TL
    3) If scan masuk exists but late:
       - TL1/TL2/TL3 depending late minutes
       - Can become H if overtime >= late (policy)
    4) If pulang < batas -> K8
    5) Else -> H
    """

    # --- Manual override first
    manual = str(row.get("Manual Status", "")).strip().upper()
    if manual in MANUAL_ALLOWED:
        return manual, 0, 0

    # --- Scan inputs
    scan_masuk = parse_time_safe(row.get("Scan Masuk"))
    scan_pulang = parse_time_safe(row.get("Scan Pulang"))

    batas_masuk_t = parse_time_safe(batas_masuk)
    batas_pulang_t = parse_time_safe(batas_pulang)

    # --- Missing scan means TL (Tidak Lengkap)
    if scan_masuk is None or scan_pulang is None:
        return "TL", 0, 0

    # --- Late calculation
    telat = minutes_diff(batas_masuk_t, scan_masuk)
    telat = telat if telat and telat > 0 else 0

    # --- Early leave
    if scan_pulang < batas_pulang_t:
        return "K8", telat, 0

    # --- Overtime
    overtime = minutes_diff(batas_pulang_t, scan_pulang)
    overtime = overtime if overtime and overtime > 0 else 0

    # --- Not late
    if telat == 0:
        return "H", 0, overtime

    # --- Late rules
    if 1 <= telat <= 30:
        return ("H" if overtime >= telat else "TL1"), telat, overtime
    if 31 <= telat <= 60:
        return ("H" if overtime >= telat else "TL2"), telat, overtime
    if telat >= 61:
        return "TL3", telat, overtime

    return "TL", telat, overtime


# =========================================================
# Matrix Builder
# =========================================================
def build_matrix(df: pd.DataFrame, periode_text: str, batas_masuk: str, batas_pulang: str) -> pd.DataFrame:
    df = df.copy()

    # enforce datetime
    df["Tanggal"] = pd.to_datetime(df["Tanggal"], errors="coerce")
    df = df.dropna(subset=["Tanggal"])
    df["Day"] = df["Tanggal"].dt.day

    # ensure columns
    for col in ["Nama", "NIK", "Unit", "GOL", "Manual Status"]:
        if col not in df.columns:
            df[col] = ""

    # Final status
    df["Status Final"] = df.apply(
        lambda r: compute_daily_status(r, batas_masuk, batas_pulang)[0],
        axis=1,
    )

    # pivot: one row per employee
    days = list(range(1, 32))
    pivot = df.pivot_table(
        index=["Nama", "NIK", "Unit", "GOL"],
        columns="Day",
        values="Status Final",
        aggfunc="last",
        fill_value="",
    )

    for d in days:
        if d not in pivot.columns:
            pivot[d] = ""
    pivot = pivot[days]

    # totals
    totals = pd.DataFrame(index=pivot.index)
    totals["HADIR"] = (pivot == "H").sum(axis=1)
    totals["SAKIT"] = (pivot == "S").sum(axis=1)
    totals["IZIN"] = (pivot == "I").sum(axis=1)
    totals["CUTI"] = (pivot == "C").sum(axis=1)
    totals["DL"] = (pivot == "DL").sum(axis=1)
    totals["K8"] = (pivot == "K8").sum(axis=1)
    totals["TL"] = (pivot == "TL").sum(axis=1)
    totals["TL1"] = (pivot == "TL1").sum(axis=1)
    totals["TL2"] = (pivot == "TL2").sum(axis=1)
    totals["TL3"] = (pivot == "TL3").sum(axis=1)
    totals["ALPA"] = (pivot == "").sum(axis=1)
    totals["JUMLAH HARI"] = (pivot != "").sum(axis=1)

    out = pivot.join(totals).reset_index()
    out.insert(0, "NO", range(1, len(out) + 1))
    return out


def build_employee_matrix_row(df: pd.DataFrame, nama: str, batas_masuk: str, batas_pulang: str) -> pd.DataFrame:
    df_emp = df[df["Nama"] == nama].copy()
    if df_emp.empty:
        return pd.DataFrame()
    mx = build_matrix(df_emp, "periode", batas_masuk, batas_pulang)
    return mx


# =========================================================
# Export Excel (Premium)
# =========================================================
FILL_MAP = {
    "H": PatternFill("solid", fgColor="C6EFCE"),
    "K8": PatternFill("solid", fgColor="FCE4D6"),
    "TL1": PatternFill("solid", fgColor="FFEB9C"),
    "TL2": PatternFill("solid", fgColor="FFD966"),
    "TL3": PatternFill("solid", fgColor="F8CBAD"),
    "TL": PatternFill("solid", fgColor="D9D9D9"),
    "S": PatternFill("solid", fgColor="BDD7EE"),
    "I": PatternFill("solid", fgColor="C9DAF8"),
    "C": PatternFill("solid", fgColor="D9E1F2"),
    "DL": PatternFill("solid", fgColor="B4C6E7"),
    "": PatternFill("solid", fgColor="FFFFFF"),
}


def save_matrix_excel(matrix_df: pd.DataFrame, out_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIX"

    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(matrix_df, index=False, header=True), start=1):
        ws.append(row)
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                v = str(cell.value) if cell.value is not None else ""
                if v in FILL_MAP:
                    cell.fill = FILL_MAP[v]

    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 22

    wb.save(out_path)
    return out_path


# =========================================================
# Public API (used by pages/generate.py)
# =========================================================
def generate_matrix_reports(
    df: pd.DataFrame,
    outdir: str,
    periode: str,
    batas_masuk: str,
    batas_pulang: str,
    manual_override_df: Optional[pd.DataFrame] = None,
):
    os.makedirs(outdir, exist_ok=True)

    df = df.copy()
    df["Tanggal"] = pd.to_datetime(df["Tanggal"], dayfirst=True, errors="coerce")

    # ✅ Apply override before matrix generation
    df = apply_manual_override(df, manual_override_df)

    matrix_all = build_matrix(df, periode, batas_masuk, batas_pulang)

    matrix_path = os.path.join(outdir, f"MATRIX_REKAP_{safe_folder_name(periode)}.xlsx")
    save_matrix_excel(matrix_all, matrix_path)

    # Multi sheet matrix per pegawai
    per_path = os.path.join(outdir, f"MATRIX_PER_PEGAWAI_{safe_folder_name(periode)}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    for nama in matrix_all["Nama"].dropna().unique():
        sub = matrix_all[matrix_all["Nama"] == nama].copy()
        ws = wb.create_sheet(title=str(nama)[:30])
        for row in dataframe_to_rows(sub, index=False, header=True):
            ws.append(row)

    wb.save(per_path)
    return matrix_path, per_path, matrix_all


# =========================================================
# ReportLab Matrix Table for PDF
# =========================================================
def matrix_to_reportlab_table(employee_matrix_row: pd.DataFrame):
    if employee_matrix_row is None or employee_matrix_row.empty:
        return None

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("mx_title", parent=styles["Heading2"], fontSize=13, alignment=1)
    subtitle_style = ParagraphStyle(
        "mx_sub",
        parent=styles["Normal"],
        fontSize=9,
        alignment=1,
        textColor=colors.HexColor("#555555"),
    )

    row = employee_matrix_row.iloc[0].to_dict()
    days = [str(i) for i in range(1, 32)]
    status_values = [str(row.get(i, "")) for i in range(1, 32)]

    table_header = ["Hari"] + days
    table_data = [table_header, ["Status"] + status_values]

    tbl = Table(table_data, colWidths=[1.5 * cm] + [0.55 * cm] * 31)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D6DCE6")),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
    ]))

    fill_map = {
        "H": colors.HexColor("#C6EFCE"),
        "K8": colors.HexColor("#FCE4D6"),
        "TL1": colors.HexColor("#FFEB9C"),
        "TL2": colors.HexColor("#FFD966"),
        "TL3": colors.HexColor("#F8CBAD"),
        "TL": colors.HexColor("#D9D9D9"),
        "S": colors.HexColor("#BDD7EE"),
        "I": colors.HexColor("#C9DAF8"),
        "C": colors.HexColor("#D9E1F2"),
        "DL": colors.HexColor("#B4C6E7"),
        "": colors.white,
    }

    for i, val in enumerate(status_values):
        col = i + 1
        tbl.setStyle(TableStyle([("BACKGROUND", (col, 1), (col, 1), fill_map.get(val, colors.white))]))

    legend = Paragraph(
        "<b>Legenda:</b> H=Hadir | K8=Kurang 8 Jam | TL=Tidak Lengkap | "
        "TL1=Telat 1-30 | TL2=Telat 31-60 | TL3=Telat ≥61 | "
        "S=Sakit | I=Izin | C=Cuti | DL=Dinas Luar",
        subtitle_style,
    )

    title = Paragraph("Matriks Kehadiran Harian", title_style)
    return [title, Spacer(1, 0.2 * cm), tbl, Spacer(1, 0.2 * cm), legend]
